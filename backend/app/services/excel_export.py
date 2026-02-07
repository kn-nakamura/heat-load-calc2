from __future__ import annotations

import io
import json
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

try:
    import yaml  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    yaml = None

from app.models.schemas import CalcResult, Project


def _resolve_value(data: dict, path: str):
    current = data
    for part in path.split("."):
        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None
    return current


def _find_sheet_by_trimmed_name(wb: openpyxl.Workbook, name: str):
    for ws in wb.worksheets:
        if ws.title.strip() == name.strip():
            return ws
    return None


def _resolve_writable_cell(ws: openpyxl.worksheet.worksheet.Worksheet, cell_ref: str):
    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        return cell
    row = cell.row
    col = cell.column
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(row=merged.min_row, column=merged.min_col)
    return ws[cell_ref]


def _load_config(yaml_path: Path, json_path: Path) -> dict:
    if yaml is not None and yaml_path.exists():
        with yaml_path.open("r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    if json_path.exists():
        with json_path.open("r", encoding="utf-8") as f:
            return json.load(f) or {}
    return {}


def export_excel(project: Project, calc_result: CalcResult | None = None) -> bytes:
    backend_dir = Path(__file__).resolve().parents[2]
    root_dir = backend_dir.parent

    template_path = root_dir / "熱負荷計算書テンプレート.xlsx"
    mapping_path = backend_dir / "app" / "config" / "excel_mapping.yml"
    mapping_json_path = backend_dir / "app" / "config" / "excel_mapping.json"
    injection_path = backend_dir / "app" / "config" / "excel_formula_injection.yml"
    injection_json_path = backend_dir / "app" / "config" / "excel_formula_injection.json"

    wb = openpyxl.load_workbook(template_path)
    mapping_cfg = _load_config(mapping_path, mapping_json_path)

    context = {
        "project": project.model_dump(),
        "calc": calc_result.model_dump() if calc_result is not None else {},
    }

    for sheet_name, entries in (mapping_cfg.get("sheets") or {}).items():
        ws = _find_sheet_by_trimmed_name(wb, sheet_name)
        if ws is None:
            continue
        for entry in entries:
            cell = entry.get("cell")
            source = entry.get("source")
            if not cell or not source:
                continue
            value = _resolve_value(context, source)
            if value is not None:
                target = _resolve_writable_cell(ws, cell)
                target.value = value

    inj_cfg = _load_config(injection_path, injection_json_path)

    for inj in inj_cfg.get("injections", []):
        ws = _find_sheet_by_trimmed_name(wb, inj["sheet"])
        if ws is None:
            continue
        cell = ws[inj["cell"]]
        mode = inj.get("mode", "if_missing")
        if mode == "if_missing" and isinstance(cell.value, str) and cell.value.startswith("="):
            continue
        if mode == "if_blank" and cell.value not in (None, ""):
            continue
        cell.value = inj["formula"]

    wb.calculation.fullCalcOnLoad = True

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
